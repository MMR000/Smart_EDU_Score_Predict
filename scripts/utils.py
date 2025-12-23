import os
import re
import json
import math
import random
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

def ensure_dir(p: str):
    os.makedirs(p, exist_ok=True)

def set_seed(seed: int):
    random.seed(seed)
    np.random.seed(seed)
    try:
        import torch
        torch.manual_seed(seed)
        if torch.cuda.is_available():
            torch.cuda.manual_seed_all(seed)
    except Exception:
        pass

def load_yaml(path: str) -> dict:
    import yaml
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)

def norm_col(s: str) -> str:
    s = str(s).strip().lower()
    s = re.sub(r"[\s\-_/]+", "", s)
    s = re.sub(r"[^a-z0-9]+", "", s)
    return s

def best_match(headers: List[str], candidates: List[str]) -> Optional[str]:
    # exact normalized match first; then substring match
    h_norm = {norm_col(h): h for h in headers}
    cand_norm = [norm_col(c) for c in candidates]
    for c in cand_norm:
        if c in h_norm:
            return h_norm[c]
    # substring heuristic
    for c in cand_norm:
        for hn, orig in h_norm.items():
            if c and (c in hn or hn in c):
                return orig
    return None

def parse_start_year(ay: str) -> int:
    m = re.search(r"(\d{4})", str(ay))
    return int(m.group(1)) if m else -1

def safe_float(x):
    return pd.to_numeric(x, errors="coerce")

def choose_sheet_by_columns(xlsx_path: str, needed: List[str], sheet_hint: Optional[str] = None) -> str:
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_hint and sheet_hint in wb.sheetnames:
        return sheet_hint

    needed_norm = [norm_col(x) for x in needed]
    best = None
    best_score = -1
    for name in wb.sheetnames:
        ws = wb[name]
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header:
            continue
        header = [str(x).strip() if x is not None else "" for x in header]
        hset = set([norm_col(h) for h in header if h])
        score = sum([1 for n in needed_norm if n in hset])
        if score > best_score:
            best_score = score
            best = name
    if best is None:
        raise ValueError(f"Cannot detect sheet for {xlsx_path}")
    return best

def read_excel_auto(xlsx_path: str, sheet: Optional[str] = None) -> pd.DataFrame:
    return pd.read_excel(xlsx_path, sheet_name=sheet, engine="openpyxl")

def detect_columns(df: pd.DataFrame) -> Dict[str, str]:
    headers = list(df.columns)

    # main grade table
    student_id = best_match(headers, ["StudentID", "student_id", "student id", "id", "studentnumber", "Student Number", "Student Code"])
    subject = best_match(headers, ["Subject", "Discipline", "Course", "CourseName", "Название дисциплины"])
    mark = best_match(headers, ["Mark", "Grade", "Score", "Result", "Баллы", "Оценка"])
    term = best_match(headers, ["Term", "Semester", "Semestr", "семестр"])
    academic_year = best_match(headers, ["Academic Year", "AcademicYear", "Year", "Study Year", "Учебный год", "Год"])

    out = {}
    for k, v in [("student_id", student_id), ("subject", subject), ("mark", mark), ("term", term), ("academic_year", academic_year)]:
        if v is not None:
            out[k] = v
    return out

def markdown_table(df: pd.DataFrame) -> str:
    try:
        return df.to_markdown(index=False)
    except Exception:
        return df.to_csv(index=False)

def save_latex_table(df: pd.DataFrame, path: str):
    ensure_dir(os.path.dirname(path))
    try:
        tex = df.to_latex(index=False, escape=True)
    except Exception:
        # fallback simple
        tex = df.to_csv(index=False)
    with open(path, "w", encoding="utf-8") as f:
        f.write(tex)
